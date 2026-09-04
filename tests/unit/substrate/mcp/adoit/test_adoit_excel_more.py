"""src/lab/substrate/mcp/adoit/adoit_excel.generate() — the branches tests/unit/substrate/mcp/adoit/test_adoit_excel.py leaves open, against
a SYNTHETIC template built with openpyxl (so each branch is deterministic, not template-dependent):
missing template, bare-list specs, elements without type/name/id, a sheet with no Name column, a
pre-filled first data row, an icon type whose sheet lacks the `Display as icon` column, relations
whose column does not exist, and duplicate relation targets. Offline.
Run: .venv/bin/python tests/unit/substrate/mcp/adoit/test_adoit_excel_more.py   (also pytest-compatible)"""
import os
import tempfile


import openpyxl

from lab.substrate.mcp.adoit import adoit_excel

TMP = tempfile.mkdtemp(prefix="adoit-excel-more-")


def _template(path):
    """A tiny ADOIT-shaped template: one sheet per type, header row, relation columns."""
    wb = openpyxl.Workbook()
    ac = wb.active; ac.title = "Application Component"
    ac.append(["Name (simple)", "Description (simple)", "Composition (->Application Component)",
               "Access (->Data Object)", "Display as icon (bool)"])
    ac.append(["Legacy App", "already in the template", None, None, None])     # pre-filled first data row
    ai = wb.create_sheet("Application Interface")
    ai.append(["Name (simple)", "Description (simple)"])                        # icon type, no icon column
    do = wb.create_sheet("Data Object")
    do.append(["Name (simple)", "Beschreibung", "Association (->Data Object)"])
    nd = wb.create_sheet("Node")
    nd.append(["ID", "Foo"])                                                    # no Name column
    wb.save(path)
    return path


def _row(ws, r):
    return {str(c.value or "").strip(): ws.cell(row=r, column=i + 1).value for i, c in enumerate(ws[1]) if c.value}


def test_missing_template_is_an_error():
    try:
        adoit_excel.generate({"elements": []}, os.path.join(TMP, "x.xlsx"), template_path=os.path.join(TMP, "none.xlsx"))
    except FileNotFoundError as e:
        assert "template missing" in str(e)
    else:
        raise AssertionError("expected FileNotFoundError")


def test_bare_list_spec_and_helpers():
    els = [{"id": "a", "type": "ApplicationComponent", "name": "A"}]
    assert adoit_excel._elements(els) is els and adoit_excel._elements({"elements": None}) == []
    assert adoit_excel._relations(els) == [] and adoit_excel._relations({}) == []
    assert adoit_excel._relations({"relationships": [{"type": "Flow"}]}) == [{"type": "Flow"}]
    assert adoit_excel._find_col(["ID", "Foo"], "Name") is None
    tpl = _template(os.path.join(TMP, "t1.xlsx"))
    out = os.path.join(TMP, "list.xlsx")
    res = adoit_excel.generate(els, out, template_path=tpl)
    assert res["objects"] == 1 and res["relations"] == 0 and res["sheets"] == ["Application Component"]
    ws = openpyxl.load_workbook(out)["Application Component"]
    assert _row(ws, 2)["Name (simple)"] == "Legacy App"                        # the template row is kept
    assert _row(ws, 3)["Name (simple)"] == "A" and _row(ws, 3)["Display as icon (bool)"] == "true"


def test_elements_without_type_name_or_id_and_sheet_without_name_column():
    tpl = _template(os.path.join(TMP, "t2.xlsx"))
    spec = {"elements": [
        {"id": "x", "name": "no type"},                                        # skipped silently
        {"id": "y", "type": "ApplicationComponent"},                           # no name: skipped silently
        {"type": "ApplicationComponent", "name": "Anonymous", "doc": "no id"},  # written, no row_of entry
        {"id": "n1", "type": "Node", "name": "Server"},                        # sheet has no Name column
        {"id": "if1", "type": "ApplicationInterface", "name": "API"},          # icon type, no icon column
        {"id": "d1", "type": "DataObject", "name": "Claim", "doc": "German header"},
    ], "relations": [{"type": "Serving", "src": "if1", "tgt": "d1"}]}
    out = os.path.join(TMP, "edge.xlsx")
    res = adoit_excel.generate(spec, out, template_path=tpl)
    assert res["objects"] == 3
    assert res["skipped"] == [{"name": "Server", "type": "Node"}]
    assert any("has no Name column" in w and "'Server'" in w for w in res["warnings"])
    assert res["sheets"] == ["Application Component", "Application Interface", "Data Object"]
    wb = openpyxl.load_workbook(out)
    assert _row(wb["Application Component"], 3)["Name (simple)"] == "Anonymous"
    assert _row(wb["Application Interface"], 2) == {"Name (simple)": "API", "Description (simple)": None}
    assert _row(wb["Data Object"], 2)["Beschreibung"] == "German header"
    assert res["relations"] == 0 and res["relations_skipped"] == 1             # no 'Serving (->Data Object)' column


def test_relations_missing_column_duplicates_and_aliases():
    tpl = _template(os.path.join(TMP, "t3.xlsx"))
    spec = {"elements": [
        {"id": "a", "type": "ApplicationComponent", "name": "A"},
        {"id": "b", "type": "ApplicationComponent", "name": "B"},
        {"id": "c", "type": "ApplicationComponent", "name": "C"},
        {"id": "d", "type": "DataObject", "name": "D"},
        {"id": "z", "type": "Nonsense", "name": "Z"},
    ], "relations": [
        {"type": "Composition", "src": "a", "tgt": "b"},
        {"type": "Composition", "src": "a", "tgt": "c"},                     # ';'-joined on the same cell
        {"type": "Composition", "src": "a", "tgt": "b"},                     # duplicate target: not re-added
        {"type": "Realisation", "src": "a", "tgt": "d"},                     # alias -> 'Realization', no column
        {"type": "Composition", "from": "b", "to": "c"},                     # from/to shape tolerated
        {"type": "Composition", "src": "a", "tgt": "z"},                     # target has no sheet
        {"type": "Composition", "src": "z", "tgt": "a"},                     # source never written
        {"src": "a", "tgt": "b"},                                            # no type
    ]}
    out = os.path.join(TMP, "rels.xlsx")
    res = adoit_excel.generate(spec, out, template_path=tpl)
    assert res["objects"] == 4 and res["skipped"] == [{"name": "Z", "type": "Nonsense"}]
    assert res["relations"] == 4 and res["relations_skipped"] == 4
    assert any(w.startswith("no column 'Realization (->Data Object)'") for w in res["warnings"])
    ws = openpyxl.load_workbook(out)["Application Component"]
    a, b = _row(ws, 3), _row(ws, 4)
    assert a["Composition (->Application Component)"] == "B;C"
    assert b["Composition (->Application Component)"] == "C"


def test_icon_types_override_empty_set():
    tpl = _template(os.path.join(TMP, "t4.xlsx"))
    out = os.path.join(TMP, "noicon.xlsx")
    adoit_excel.generate({"elements": [{"id": "a", "type": "ApplicationComponent", "name": "A"}]}, out,
                         template_path=tpl, icon_types=set())
    assert _row(openpyxl.load_workbook(out)["Application Component"], 3)["Display as icon (bool)"] is None


if __name__ == "__main__":
    for name, fn in list(globals().items()):
        if name.startswith("test_") and callable(fn):
            fn(); print("ok", name)
