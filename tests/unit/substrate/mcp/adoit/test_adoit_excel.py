"""src/lab/substrate/mcp/adoit/adoit_excel.generate() over a small engine spec into a temp copy of the bundled
ADOIT template: sheet chosen per type (normalised match), Name/Description cells, the
`Composition (->Application Component)` relation column on the SOURCE row, `Display as icon`.
Offline: openpyxl + the vendored template only.
Run: .venv/bin/python tests/unit/substrate/mcp/adoit/test_adoit_excel.py   (also pytest-compatible)"""
import tempfile
from pathlib import Path


import openpyxl

from lab.substrate.mcp.adoit import adoit_excel

SPEC = {
    "name": "Claims Portal", "id": "claims-portal",
    "elements": [
        {"id": "portal", "type": "ApplicationComponent", "name": "Portal", "doc": "Web front end"},
        {"id": "adjudication", "type": "ApplicationComponent", "name": "Adjudication Module", "doc": "Decides claims"},
        {"id": "claim", "type": "DataObject", "name": "Claim", "doc": "A filed claim record"},
    ],
    "relations": [
        {"id": "r1", "type": "Composition", "src": "portal", "tgt": "adjudication"},
        {"id": "r2", "type": "Access", "src": "adjudication", "tgt": "claim", "accessType": "Write"},
    ],
}


def _headers(ws):
    return [str(c.value or "").strip() for c in ws[1]]


def _row(ws, r):
    hdr = _headers(ws)
    return {h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(hdr) if h}


def test_generate_into_template():
    with tempfile.TemporaryDirectory() as td:
        out = str(Path(td) / "import.xlsx")
        res = adoit_excel.generate(SPEC, out)
        assert res["path"] == out and res["objects"] == 3 and res["skipped"] == [] and res["warnings"] == []
        assert res["relations"] == 2 and res["relations_skipped"] == 0
        assert res["sheets"] == ["Application Component", "Data Object"]   # squash('ApplicationComponent') -> sheet

        wb = openpyxl.load_workbook(out)
        ac = wb["Application Component"]
        portal, adjudication = _row(ac, 2), _row(ac, 3)
        assert portal["Name (simple)"] == "Portal" and portal["Description (simple)"] == "Web front end"
        assert adjudication["Name (simple)"] == "Adjudication Module"
        # relation on the SOURCE row, column '<Relation> (->TargetSheet)', value = target NAME
        assert portal["Composition (->Application Component)"] == "Adjudication Module"
        assert adjudication["Composition (->Application Component)"] is None
        assert adjudication["Access (->Data Object)"] == "Claim"
        # representation pass: components get the compact icon glyph
        assert portal["Display as icon (bool)"] == "true" and adjudication["Display as icon (bool)"] == "true"
        do = wb["Data Object"]
        claim = _row(do, 2)
        assert claim["Name (simple)"] == "Claim" and claim["Description (simple)"] == "A filed claim record"
        assert claim.get("Display as icon (bool)") in (None, "")               # not an ICON_TYPE
        assert all(v in (None, "") for v in _row(ac, 4).values())              # nothing beyond the 2 rows


def test_unknown_type_and_missing_column_are_reported_not_raised():
    spec = {"elements": [{"id": "a", "type": "Nonsense", "name": "A"},
                         {"id": "b", "type": "ApplicationComponent", "name": "B"},
                         {"id": "c", "type": "ApplicationComponent", "name": "C"}],
            "relations": [{"type": "Flow", "src": "b", "tgt": "c"},          # no 'Flow (->…)' column? -> skipped+warned or written
                          {"type": "Serving", "src": "a", "tgt": "b"}]}       # source not written -> skipped silently
    with tempfile.TemporaryDirectory() as td:
        res = adoit_excel.generate(spec, str(Path(td) / "x.xlsx"))
        assert res["objects"] == 2 and res["skipped"] == [{"name": "A", "type": "Nonsense"}]
        assert any("no template sheet for type 'Nonsense'" in w for w in res["warnings"])
        assert res["relations"] + res["relations_skipped"] == 2 and res["relations_skipped"] >= 1


def test_icon_types_override_and_british_alias():
    spec = {"elements": [{"id": "s", "type": "ApplicationService", "name": "S"},
                         {"id": "c", "type": "ApplicationComponent", "name": "C"}],
            "relations": [{"type": "Realisation", "src": "c", "tgt": "s"}]}  # British spelling -> template column
    with tempfile.TemporaryDirectory() as td:
        out = str(Path(td) / "y.xlsx")
        res = adoit_excel.generate(spec, out, icon_types={"application service"})   # squash-matched membership
        assert res["relations"] == 1 and res["relations_skipped"] == 0
        wb = openpyxl.load_workbook(out)
        assert _row(wb["Application Service"], 2)["Display as icon (bool)"] == "true"
        c = _row(wb["Application Component"], 2)
        assert c["Display as icon (bool)"] in (None, "") and c["Realization (->Application Service)"] == "S"


if __name__ == "__main__":
    test_generate_into_template()
    test_unknown_type_and_missing_column_are_reported_not_raised()
    test_icon_types_override_and_british_alias()
    print("ALL TESTS PASSED")
