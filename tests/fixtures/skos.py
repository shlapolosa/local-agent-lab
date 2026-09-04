"""Shared doubles/harness hoisted from the former `test_skos` module (restructure): imported by every test that
needs them (`from fixtures.skos import …`) instead of test-to-test imports.
"""
import tempfile

import openpyxl
from rdflib import RDF, RDFS, Literal, URIRef

from lab.core.semantic.ontology import META, Ontology
from lab.core.semantic.reference import baguild
from lab.core.semantic.skos import KIND_TO_ARCHIMATE, SKOS, SkosScheme, concept_id


BASE = "urn:lab:semantic:ref:syn#"


def scheme(source="fixture"):
    """Two capability trees (one 3 levels deep), a value stream, two related information concepts."""
    return SkosScheme("syn-v1", BASE, "Synthetic Model", [
        {"id": "c1", "label": "Care Delivery", "definition": "Deliver care", "kind": "capability", "parent": None, "level": 1, "tier": 1},
        {"id": "c11", "label": "Triage", "definition": "", "kind": "capability", "parent": "c1", "level": 2, "tier": 1},
        {"id": "c111", "label": "Urgent Triage", "definition": "Fast lane", "kind": "capability", "parent": "c11", "level": 3, "tier": 1},
        {"id": "c12", "label": "Discharge", "definition": None, "kind": "capability", "parent": "c1", "level": 2, "tier": 2},
        {"id": "c2", "label": "Billing", "definition": "Bill", "kind": "capability", "parent": None, "level": 1, "tier": 3},
        {"id": "vs1", "label": "Admit Patient", "definition": "Arrive to bed", "kind": "value-stream", "parent": None, "level": 1, "tier": None},
        {"id": "i1", "label": "Patient", "definition": "", "kind": "information", "parent": None, "level": 1, "tier": None, "related": ["i2"]},
        {"id": "i2", "label": "Encounter", "definition": "", "kind": "information", "parent": None, "level": 1, "tier": None, "related": []},
    ], source=source)


# ---------------------------------------------------------------- synthetic workbook authoring
def write_workbook(path, caps=(), value_streams=(), org=(), stakeholders=(), info=(), sheets=None):
    """caps: (tier, level, label, definition); value_streams: (name, definition);
    org: (level, label, _, definition); stakeholders: (type, category, name, definition);
    info: (concept, _, definition, types, related, states). Only the sheets with rows are written,
    unless `sheets` forces a set of (possibly empty) sheet names."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    titled = {"Capability Map": (("Capability Map", None, None, None), ("Tier", "Level", "Capability", "Definition"), caps),
              "Value Stream Inventory": (("Value Stream Inventory", None), ("Value Stream Name", "Definition"), value_streams),
              "Organization Map": (("Organization Map", None, None, None), ("Business Unit Level", "Business Unit", "x", "Definition"), org),
              "Stakeholder Map": (("Stakeholder Map", None, None, None), ("Stakeholder Type", "Category", "Stakeholder", "Definition"), stakeholders),
              "Information Map": (("Information Map",), ("Information Concept", "x", "Definition", "Types", "Related Information Concepts", "States"), info)}
    for name, (title, header, rows) in titled.items():
        if rows or (sheets and name in sheets):
            ws = wb.create_sheet(name)
            ws.append(title); ws.append(header); ws.append([None] * len(header))       # blank row is skipped
            for r in rows:
                ws.append(list(r))
    if not wb.sheetnames:
        wb.create_sheet("Notes")
    wb.save(path)
    return path


HEALTH = [(1, 1, "Care Delivery", "Deliver care"), (1, 2, "Triage", "Sort"), (1, 3, "Urgent Triage", ""),
          (1, 2, "Discharge", "Out"), (2, 1, "Billing", "Bill"), (2, 2, "", "no label -> skipped"),
          ("Tier", "x", "junk", "non-numeric level -> skipped"), (2, 2, "Claims", "")]
INSURE = [(1, 1, "Billing", "Bill (insurance)"), (1, 2, "Premium Collection", ""), (3, 1, "Underwriting", "")]
