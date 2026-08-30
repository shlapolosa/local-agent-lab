"""Business Architecture Guild reference models (origin workbooks) -> SkosScheme.

Sheets understood: 'Capability Map' (Tier | Level | Capability | Definition; hierarchy
implicit from row order + Level), 'Value Stream(s) Inventory', and — when present —
'Stakeholder Map', 'Organization Map' (Business Unit Level hierarchy), 'Information Map'
(Related Information Concepts -> skos:related). The workbooks are licensed: they stay in
semantic/reference/sources/ (git-ignored) or REFERENCE_MODELS_DIR; only derived RDF is built.
"""
import glob
import os

import openpyxl

from ..skos import SkosScheme, concept_id

HERE = os.path.dirname(__file__)


def _rows(ws):
    for r in ws.iter_rows(values_only=True):
        if r and any(c is not None and str(c).strip() for c in r):
            yield [str(c).strip() if c is not None else "" for c in r]


def _sheet(wb, *names):
    for n in wb.sheetnames:
        if n.strip().lower() in {x.lower() for x in names}:
            return wb[n]
    return None


def _hierarchy(rows, scheme, kind, level_col, label_col, def_col, tier_col=None, path_prefix=()):
    """Rows with a numeric level column -> concepts with parents from a level stack."""
    out, stack = [], []          # stack: [(level, id, label)]
    for r in rows:
        try:
            lvl = int(float(r[level_col]))
        except (ValueError, IndexError):
            continue
        label = r[label_col]
        if not label:
            continue
        while stack and stack[-1][0] >= lvl:
            stack.pop()
        path = tuple(path_prefix) + tuple(s[2] for s in stack) + (label,)
        cid = concept_id(scheme, path)
        out.append({"id": cid, "label": label, "definition": r[def_col] if def_col is not None and len(r) > def_col else "",
                    "kind": kind, "parent": stack[-1][1] if stack else None, "level": lvl,
                    "tier": (int(float(r[tier_col])) if tier_col is not None and r[tier_col] else None)})
        stack.append((lvl, cid, label))
    return out


def parse(path, name, title):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    concepts = []
    ws = _sheet(wb, "Capability Map")
    if ws is not None:
        rows = [r for r in _rows(ws) if r[0] not in ("Capability Map", "Tier")]
        concepts += _hierarchy(rows, name, "capability", level_col=1, label_col=2, def_col=3, tier_col=0)
    ws = _sheet(wb, "Value Stream Inventory", "Value Streams Inventory")
    if ws is not None:
        for r in _rows(ws):
            if r[0] in ("Value Stream Inventory", "Value Stream Name") or not r[0]:
                continue
            concepts.append({"id": concept_id(name, ("value-stream", r[0])), "label": r[0],
                             "definition": r[1] if len(r) > 1 else "", "kind": "value-stream",
                             "parent": None, "level": 1, "tier": None})
    ws = _sheet(wb, "Organization Map")
    if ws is not None:
        rows = [r for r in _rows(ws) if r[0] not in ("Organization Map", "Business Unit Level")]
        concepts += _hierarchy(rows, name, "org-unit", level_col=0, label_col=1, def_col=3, path_prefix=("org",))
    ws = _sheet(wb, "Stakeholder Map")
    if ws is not None:
        cats = {}
        for r in _rows(ws):
            if r[0] in ("Stakeholder Map", "Stakeholder Type") or len(r) < 3 or not r[2]:
                continue
            cat_key = (r[0], r[1])
            if cat_key not in cats:
                cid = concept_id(name, ("stakeholder", r[0], r[1]))
                cats[cat_key] = cid
                concepts.append({"id": cid, "label": f"{r[1]} ({r[0]})", "definition": "", "kind": "stakeholder",
                                 "parent": None, "level": 1, "tier": None})
            concepts.append({"id": concept_id(name, ("stakeholder", r[0], r[1], r[2])), "label": r[2],
                             "definition": r[3] if len(r) > 3 else "", "kind": "stakeholder",
                             "parent": cats[cat_key], "level": 2, "tier": None})
    ws = _sheet(wb, "Information Map")
    if ws is not None:
        info, related = [], {}
        for r in _rows(ws):
            if r[0] in ("Information Map", "Information Concept") or not r[0]:
                continue
            cid = concept_id(name, ("information", r[0]))
            info.append({"id": cid, "label": r[0], "definition": r[2] if len(r) > 2 else "", "kind": "information",
                         "parent": None, "level": 1, "tier": None,
                         "types": r[3] if len(r) > 3 else "", "states": r[5] if len(r) > 5 else ""})
            related[cid] = [x.strip() for x in (r[4] if len(r) > 4 else "").split(",") if x.strip()]
        by_label = {c["label"].lower(): c["id"] for c in info}
        for c in info:
            c["related"] = [by_label[x.lower()] for x in related[c["id"]] if x.lower() in by_label]
        concepts += info
    return SkosScheme(name=name, base=f"urn:lab:semantic:ref:{name}#", title=title, concepts=concepts,
                      source=os.path.basename(path))


KNOWN = {  # filename stem -> (scheme name, title)
    "healthcare-provider-v2.0": ("healthcare-provider-v2.0", "BA Guild Healthcare Provider Reference Model v2.0"),
    "insurance-v5.0": ("insurance-v5.0", "BA Guild Insurance Reference Model v5.0"),
}


def load_all(directory=None):
    d = directory or os.environ.get("REFERENCE_MODELS_DIR") or os.path.join(HERE, "sources")
    schemes = []
    for p in sorted(glob.glob(os.path.join(d, "*.xlsx"))):
        stem = os.path.splitext(os.path.basename(p))[0]
        name, title = KNOWN.get(stem, (stem, stem))
        schemes.append(parse(p, name, title))
    return schemes
