"""Generate an ADOIT **Excel object-import** file — the CE-safe write path for OBJECTS.

Why this exists: on the hosted Community Edition the REST write verbs are blocked at the edge
(POST/PATCH/DELETE -> "URL not available"), and the ArchiMate Model Exchange import *duplicates*
objects (it never matches on identifier — verified live). ADOIT's **"Import objects from Excel"**
is the one file path that both **creates and updates** repository objects, so it is our object
write path; the ArchiMate XML stays the path for **views/diagrams**.

Verified mechanism (live, Sep 2026):
- The import **matches a row to an existing object by NAME** (the key attribute), scoped to the
  chosen import group: a name found once is UPDATED in place, a name not present is CREATED, a name
  present more than once errors. So the `ID` column is a plain attribute, NOT the match key — name
  uniqueness is what matters (which is why the existing-architecture-aware step must avoid duplicates).
- The template has **one sheet per element type**, each with columns `Name` · `ID` · typed attributes
  (`Description`, enums, dates…) · relationship columns `<Relation> (->TargetType)`. Multiple
  references in one cell are `;`-separated; bools are `true`/`false`.

We fill the tenant's **downloaded template** (copy + populate) so sheet/column/config match exactly.
The bundled template is the ENGLISH one (`templates/adoit_object_import_template.xlsx`): its sheet
names ARE the ArchiMate types ("Application Component", "Course of Action") and its relationship
labels ARE the ArchiMate relation names ("Composition", "Serving", "Realization"). So the type->sheet
and relation->column maps are **derived from the template at runtime** (normalized match) rather than
hardcoded — swap the template for another locale and re-derivation still works for the sheet map;
non-English *relation labels* would need `REL_ALIAS` extended.

Objects carry Name + Description; **relationships** are written on the SOURCE object's row in the
`<Relation> (->TargetSheet)` column, value = the target's NAME (`;`-joined for several). All ArchiMate
structural relations are outgoing from the source; ADOIT-specific roles (RACI actors, Vendor,
Predecessor) are intentionally left unset.

A third **representation pass** sets the per-object `Display as icon (bool)` attribute to `"true"`
for a configurable set of types (`ICON_TYPES`, overridable via `generate(..., icon_types=...)`) so
ADOIT renders them with the compact ICON glyph (components as the component glyph, interfaces as
circles) instead of a full box. `Colour`/`Text position`/`Text alignment` are left untouched.
"""
from __future__ import annotations

import os
import shutil

import openpyxl

from lab.core.canon import squash  # type/sheet-name matching: 'ApplicationComponent' == 'Application Component'

_HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(_HERE, "templates", "adoit_object_import_template.xlsx")

# British -> template (US) spellings, so a spec relation still finds its column.
REL_ALIAS = {"Realisation": "Realization", "Specialisation": "Specialization"}

# ArchiMate types ADOIT should render with the compact ICON glyph (the per-object
# `Display as icon (bool)` attribute). Components/collaborations show their box glyph and the
# interfaces the circle glyph. Overridable per call via `generate(..., icon_types=...)`; membership
# is matched by normalised name (`squash`), so "ApplicationComponent" hits the template sheet
# "Application Component". ArchiMate has exactly three interface types — all listed here.
ICON_TYPES = {
    "ApplicationComponent",
    "ApplicationCollaboration",
    "ApplicationInterface",
    "BusinessInterface",
    "TechnologyInterface",
}


def _find_col(headers: list, *prefixes: str):
    """1-based column index of the first header starting with one of `prefixes` (case-insensitive)."""
    for i, h in enumerate(headers):
        hl = str(h or "").strip().lower()
        if any(hl.startswith(p.lower()) for p in prefixes):
            return i + 1
    return None


def _elements(spec) -> list:
    """Accept either the engine spec ({elements:[{id,type,name,doc,folder}]}) or a bare list."""
    if isinstance(spec, list):
        return spec
    return spec.get("elements", []) or []


def _relations(spec) -> list:
    """Engine spec relations: [{type, src, tgt}]. Tolerate a `relationships`/`from`/`to` shape too."""
    if isinstance(spec, dict):
        return spec.get("relations", []) or spec.get("relationships", []) or []
    return []


def generate(spec, out_path: str, template_path: str = DEFAULT_TEMPLATE,
             icon_types: set | None = None) -> dict:
    """Write an ADOIT Excel object-import file into the tenant's template. Passes:
      1. one row per element on its type's sheet, carrying Name + Description, and — for types in
         `icon_types` (default `ICON_TYPES`) whose sheet has a `Display as icon` column — the
         representation attribute `Display as icon = "true"` (compact ICON glyph);
      2. each relationship on the SOURCE object's row, in the column `<Relation> (->TargetSheet)`,
         value = the target's NAME (`;`-joined for several).
    Objects are matched on **Name** at import (create if new, update if present), so no ID is written —
    name uniqueness (the existing-aware step's job) is the contract. Returns a summary dict."""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"ADOIT Excel template missing: {template_path}")
    if icon_types is None:
        icon_types = ICON_TYPES
    icon_norms = {squash(t) for t in icon_types}   # normalised for squash-based type matching
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)

    sheet_by_norm = {squash(s): s for s in wb.sheetnames}   # 'applicationcomponent' -> 'Application Component'

    def sheet_for(atype):
        return sheet_by_norm.get(squash(atype))

    elements = _elements(spec)
    by_id = {e.get("id"): e for e in elements if e.get("id")}
    written, skipped, warnings = 0, [], []
    next_row: dict = {}
    row_of: dict = {}          # element id -> (sheet, row, headers)
    hdr_cache: dict = {}

    # ---- pass 1: objects ----
    for el in elements:
        atype, name = el.get("type"), el.get("name")
        if not (atype and name):
            continue
        sheet = sheet_for(atype)
        if not sheet:
            skipped.append({"name": name, "type": atype})
            warnings.append(f"no template sheet for type {atype!r} (element {name!r}) — skipped")
            continue
        ws = wb[sheet]
        headers = hdr_cache.get(sheet) or [c.value for c in ws[1]]
        hdr_cache[sheet] = headers
        c_name = _find_col(headers, "Name")
        c_desc = _find_col(headers, "Description", "Beschreibung")
        if c_name is None:
            warnings.append(f"sheet {sheet!r} has no Name column — skipped {name!r}")
            skipped.append({"name": name, "type": atype})
            continue
        if sheet not in next_row:                       # first empty data row per sheet
            r = 2
            while any(ws.cell(row=r, column=i + 1).value not in (None, "") for i in range(len(headers))):
                r += 1
            next_row[sheet] = r
        r = next_row[sheet]
        ws.cell(row=r, column=c_name).value = name
        if c_desc is not None and el.get("doc"):
            ws.cell(row=r, column=c_desc).value = el["doc"]
        # representation pass: compact ICON glyph for configured types (bool = "true"/"false"),
        # only when this sheet actually carries the attribute column.
        if squash(atype) in icon_norms:
            c_icon = _find_col(headers, "Display as icon")
            if c_icon is not None:
                ws.cell(row=r, column=c_icon).value = "true"
        if el.get("id"):
            row_of[el["id"]] = (sheet, r, headers)
        next_row[sheet] = r + 1
        written += 1

    # ---- pass 2: relationships (on the source object's row) ----
    rel_written, rel_skipped = 0, 0
    for rel in _relations(spec):
        rtype = rel.get("type")
        rtype = REL_ALIAS.get(rtype, rtype)
        src, tgt = rel.get("src") or rel.get("from"), rel.get("tgt") or rel.get("to")
        loc, tgt_el = row_of.get(src), by_id.get(tgt)
        tgt_sheet = sheet_for(tgt_el.get("type")) if tgt_el else None
        if not (rtype and loc and tgt_el and tgt_sheet):
            rel_skipped += 1
            continue
        sheet, r, headers = loc
        col_header = f"{rtype} (->{tgt_sheet})"
        ci = next((i + 1 for i, h in enumerate(headers) if str(h or "").strip() == col_header), None)
        if ci is None:                                  # ADOIT allows no such (relation, target) here
            rel_skipped += 1
            warnings.append(f"no column {col_header!r} on sheet {sheet!r} "
                            f"({rtype} -> {tgt_el.get('name')!r}) — skipped")
            continue
        ws = wb[sheet]
        cur = ws.cell(row=r, column=ci).value
        existing = [n.strip() for n in str(cur).split(";")] if cur else []
        tname = tgt_el.get("name")
        if tname not in existing:                       # multiple refs -> ';'-joined (template rule)
            ws.cell(row=r, column=ci).value = f"{cur};{tname}" if cur else tname
        rel_written += 1

    wb.save(out_path)
    return {"path": out_path, "objects": written, "sheets": sorted(next_row),
            "relations": rel_written, "relations_skipped": rel_skipped,
            "skipped": skipped, "warnings": warnings}
